import base64
import os.path
from io import BytesIO

from django import forms
from django.conf import settings
from django.contrib import admin, messages
from django.http import HttpResponse
from django.utils.html import format_html
from openpyxl import load_workbook

from .models import Client, ClientLegislation, ClientLegislationJobRole


def export(modeladmin, request, queryset):  # pylint: disable=too-many-locals

    def write_rows(_wb, _rows, target):
        rownum = 5
        sheet = _wb[target]
        for obj in _rows:
            for col, _value in enumerate(obj, 1):
                sheet.cell(row=rownum, column=col).value = _value
            rownum += 1

    def get_legislation_values(obj):
        return [
            "",  # *PwC Territory
            str(obj.identifier),  # *Unique Identifier Code
            "; ".join([_t.name for _t in obj.topic.all()]),  # *Topic
            obj.name_local,  # *Name of legislation in the original language
            obj.abbreviation,  # Abbreviation
            obj.name_generic,  # Generic name of legislation
            "; ".join([_t.name for _t in obj.issuing_jurisdiction.all()]),  # *Issuing jurisdiction
            "; ".join([_t.name for _t in obj.type.all()]),  # *Type of legislation
            "",  # Type of legislation - Other
            "; ".join([_t.name for _t in obj.geographical_scope.all()]),  # *Geographical scope
            "",  # Geographical scope - Other
            "Yes" if obj.is_in_effect else "No",  # *In effect
            obj.effective_date,  # *Applicable from
            obj.effective_until,  # *Applicable to
            obj.responsible_authority,  # *Responsible Authority,
            "; ".join([_t.name for _t in obj.product_service.all()]),  # *Relevant product (group) or services
            "",  # Relevant product (group) or services - Other
            "",  # *Type of requirements
            obj.objective,  # *Objective of the legislation (summary)
            obj.scope,  # *Scope of the legislation (summary)
            obj.responsible_party,  # *Responsible party (summary)
            "; ".join([_t.name for _t in obj.non_compliance_consequence.all()]),  # *Consequences of non-compliance
            "",  # Consequences of non-compliance - Other
            obj.link,  # *Link to legislation
            obj.additional_links,  # Link to additional guidance
            "; ".join([_t.get_full_name() for _t in obj.pwc_contact.all()]),  # *Relevant PwC contact
        ]

    def get_registration_values(obj):
        return [
            str(obj.identifier),  # *Legislation code
            obj.description,  # *General description
            obj.responsible_authority,  # *Responsible authority
            obj.trigger,  # *Activities that trigger requirements
            obj.responsible_party,  # *Responsible party
            "",  # *Data elements required
            "",  # *Payment obligations
            obj.deadline,  # *Deadline
            obj.threshold,  # *Threshold
            obj.sanctions,  # *Sanctions
            obj.exemptions,  # *Exemptions
        ]

    def get_reporting_values(obj):
        return [
            str(legislation.identifier),  # *Legislation code
            reporting.description,  # *General description
            reporting.responsible_authority,  # *Responsible authority
            reporting.trigger,  # *Activities that trigger requirements
            reporting.responsible_party,  # *Responsible party
            reporting.data_elements,  # *Data elements required
            reporting.language,  # Language of reporting
            reporting.frequency,  # Language of reporting
            reporting.way_of_submission,  # Way of submitting
            "",  # Way of submitting - Other
            reporting.payment_obligations,  # Payment obligations and rates
            reporting.retainment_of_records,  # *Retainment of records
            reporting.refund_possibilities,  # Refund possibilities
            reporting.threshold,  # Thresholds
            reporting.sanctions,  # Sanctions
            reporting.exemptions,  # Excemptions
        ]

    def get_regulatory_values(obj):
        return [
            str(legislation.identifier),  # *Legislation code
            regulatery.description,  # *General description
            regulatery.responsible_authority,  # *Responsible authority
            regulatery.trigger,  # *Activities that trigger requirements
            regulatery.responsible_party,  # *Responsible party
            regulatery.key_actions,  # Key actions for compliance
            regulatery.deadline,  # Deadlines
            regulatery.threshold,  # Thresholds
            regulatery.sanctions,  # Sanctions
            regulatery.exemptions,  # Excemptions
        ]

    workbook = load_workbook(os.path.join(settings.BASE_DIR, "core", "templates", "excel", "export_template.xlsx"))
    rows = {"legislation": [], "registration": [], "reporting": [], "regulatory": []}
    for legislation in queryset.first().legislation_list.all():
        rows["legislation"].append(get_legislation_values(legislation))
        for registration in legislation.registration_requirements.all():
            rows["registration"].append(get_registration_values(registration))
        for reporting in legislation.reporting_requirements.all():
            rows["reporting"].append(get_reporting_values(reporting))
        for regulatery in legislation.regulatory_requirements.all():
            rows["regulatory"].append(get_regulatory_values(regulatery))

    for key, value in rows.items():
        write_rows(workbook, value, key.capitalize())

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    messages.success(request, "Export ran successfully")
    response = HttpResponse(output, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=export.xlsx"

    return response


export.short_description = "Export Client Data"


class ClientAdminForm(forms.ModelForm):
    logo_clear = forms.BooleanField(required=False, label="Clear logo")

    class Meta:
        model = Client
        fields = "__all__"

    def save(self, commit=True):
        instance = super().save(commit=False)
        if self.cleaned_data.get("logo_clear"):
            instance.logo = None  # Clear the logo field
        if commit:
            instance.save()
        return instance


@admin.register(Client)
class ClientAdmin(admin.ModelAdmin):
    form = ClientAdminForm
    list_display = ("name", "is_published", "identifier", "domain", "show_logo", "sso_settings")
    exclude = ["legislation_list"]
    actions = [export]

    def show_logo(self, obj):
        if obj.logo:
            return format_html(
                '<img src="data:image/png;base64,{}" width="50" height="50" />',
                base64.b64encode(obj.logo).decode("utf-8"),
            )
        return "No Logo"

    show_logo.short_description = "Logo"


@admin.register(ClientLegislationJobRole)
class ClientLegislationJobRoleAdmin(admin.ModelAdmin):
    list_display = ("client", "legislation", "job_role", "note", "identifier")
    search_fields = ("client__name", "legislation__name", "job_role__name", "identifier", "note")
    list_filter = ("client", "legislation", "job_role")
    readonly_fields = ("identifier",)


@admin.register(ClientLegislation)
class ClientLegislationAdmin(admin.ModelAdmin):
    list_display = (
        "client",
        "legislation",
        "is_published",
        "requested_by",
        "request_date",
        "published_by",
        "publication_date",
    )
    list_filter = ("client", "legislation", "is_published")  # Filters for client, legislation, and approval status
    search_fields = ("client__name", "legislation__name")  # Enable search by client and legislation name
